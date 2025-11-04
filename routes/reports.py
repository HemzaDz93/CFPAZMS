from flask import (
    Blueprint, render_template, request, send_file, jsonify, 
    current_app, flash, redirect, url_for
)
from flask_login import login_required, current_user
from models import (
    db, Item, Transaction, User, MealRecord, 
    AssetRegistration, PurchaseOrder, OrganizationSettings, ItemCategory_Model
)
from auth_helpers import require_granular_permission
from datetime import datetime, timedelta, date
from sqlalchemy import func, and_
import io
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

reports_bp = Blueprint('reports', __name__, url_prefix='/reports')

@reports_bp.route('/')
@login_required
def index():
    """الصفحة الرئيسية للتقارير"""
    # التحقق من صلاحية عرض التقارير
    if not current_user.has_granular_permission('reports_view'):
        flash('ليس لديك صلاحية لعرض التقارير', 'danger')
        return redirect(url_for('dashboard.index'))
    
    return render_template('reports/index.html')

# ==================== Inventory Reports ====================

@reports_bp.route('/inventory-movement')
@login_required
def inventory_movement():
    """تقرير حركة المخزون"""
    # التحقق من صلاحية عرض تقارير المخزون
    if not current_user.has_granular_permission('reports_inventory_movement'):
        flash('ليس لديك صلاحية لعرض تقرير حركة المخزون', 'danger')
        return redirect(url_for('reports.index'))
    
    from_date = request.args.get('from_date')
    to_date = request.args.get('to_date')
    
    if not from_date or not to_date:
        from_date = (date.today() - timedelta(days=30)).isoformat()
        to_date = date.today().isoformat()
    
    from_date = datetime.strptime(from_date, '%Y-%m-%d')
    to_date = datetime.strptime(to_date, '%Y-%m-%d')
    to_date = to_date.replace(hour=23, minute=59, second=59)
    
    transactions = Transaction.query.filter(
        Transaction.transaction_date.between(from_date, to_date)
    ).order_by(Transaction.transaction_date).all()
    
    # احسب ملخص حسب نوع الحركة
    purchase_total = sum(t.quantity for t in transactions if t.transaction_type == 'purchase')
    purchase_count = sum(1 for t in transactions if t.transaction_type == 'purchase')
    
    issue_total = sum(t.quantity for t in transactions if t.transaction_type == 'issue')
    issue_count = sum(1 for t in transactions if t.transaction_type == 'issue')
    
    transfer_total = sum(t.quantity for t in transactions if t.transaction_type == 'transfer')
    transfer_count = sum(1 for t in transactions if t.transaction_type == 'transfer')
    
    net_balance = purchase_total - issue_total
    
    # احصل على جميع الفئات للفلتر
    categories = ItemCategory_Model.query.filter_by(is_active=True).all()
    
    org_settings = OrganizationSettings.query.first()
    
    return render_template(
        'reports/inventory_movement.html',
        transactions=transactions,
        purchase_total=purchase_total,
        purchase_count=purchase_count,
        issue_total=issue_total,
        issue_count=issue_count,
        transfer_total=transfer_total,
        transfer_count=transfer_count,
        net_balance=net_balance,
        categories=categories,
        from_date=from_date.date().isoformat(),
        to_date=to_date.date().isoformat(),
        org_settings=org_settings
    )

@reports_bp.route('/low-stock')
@login_required
def low_stock():
    """تقرير الأصناف منخفضة المخزون"""
    # التحقق من صلاحية عرض تقرير المخزون المنخفض
    if not current_user.has_granular_permission('reports_low_stock'):
        flash('ليس لديك صلاحية لعرض تقرير المخزون المنخفض', 'danger')
        return redirect(url_for('reports.index'))
    
    low_stock_items = Item.query.filter(
        Item.quantity_in_stock <= Item.minimum_quantity,
        Item.is_active == True
    ).all()
    
    # احسب الملخصات
    below_minimum = sum(1 for item in low_stock_items if item.quantity_in_stock < 0)
    warning_level = sum(1 for item in low_stock_items if item.quantity_in_stock >= 0)
    total_low_stock = len(low_stock_items)
    
    # احصل على الفئات
    categories = ItemCategory_Model.query.filter_by(is_active=True).all()
    
    org_settings = OrganizationSettings.query.first()
    
    return render_template(
        'reports/low_stock_report.html',
        items=low_stock_items,
        low_stock_items=low_stock_items,
        below_minimum=below_minimum,
        warning_level=warning_level,
        total_low_stock=total_low_stock,
        categories=categories,
        org_settings=org_settings
    )

@reports_bp.route('/asset-inventory')
@login_required
def asset_inventory():
    """جرد الأصول"""
    # التحقق من صلاحية عرض تقرير الأصول
    if not current_user.has_granular_permission('reports_asset_inventory'):
        flash('ليس لديك صلاحية لعرض تقرير الأصول', 'danger')
        return redirect(url_for('reports.index'))
    
    status = request.args.get('status', '')
    user_id = request.args.get('user_id', '')
    
    query = AssetRegistration.query
    
    if status:
        query = query.filter_by(status=status)
    
    # Import ItemIssue model
    from models import ItemIssue
    
    # جمع البيانات مع معلومات التسليم
    assets = []
    for asset in query.all():
        current_issue = None
        # البحث عن آخر تسليم للأصل (التسليم الذي لم يتم استرجاع الأصل فيه بعد)
        if asset.status != 'returned':
            current_issue = ItemIssue.query.filter(
                ItemIssue.asset_id == asset.id,
                ItemIssue.actual_return_date == None
            ).order_by(ItemIssue.issue_date.desc()).first()
        
        assets.append({
            'asset': asset,
            'current_user': current_issue.issued_to if current_issue else None
        })
    
    users = User.query.all()
    org_settings = OrganizationSettings.query.first()
    
    return render_template(
        'reports/asset_inventory.html',
        assets=assets,
        users=users,
        selected_status=status,
        org_settings=org_settings
    )

# ==================== Restaurant Reports ====================

@reports_bp.route('/meal-consumption')
@login_required
def meal_consumption():
    """تقرير استهلاك المطعم"""
    # التحقق من صلاحية عرض تقرير استهلاك المطعم
    if not current_user.has_granular_permission('reports_meal_consumption'):
        flash('ليس لديك صلاحية لعرض تقرير استهلاك المطعم', 'danger')
        return redirect(url_for('reports.index'))
    
    from_date = request.args.get('from_date')
    to_date = request.args.get('to_date')
    
    if not from_date or not to_date:
        from_date = (date.today() - timedelta(days=7)).isoformat()
        to_date = date.today().isoformat()
    
    from_date_obj = datetime.strptime(from_date, '%Y-%m-%d').date()
    to_date_obj = datetime.strptime(to_date, '%Y-%m-%d').date()
    
    meals = MealRecord.query.filter(
        MealRecord.record_date.between(from_date_obj, to_date_obj)
    ).all()
    
    # حساب الاستهلاك الإجمالي
    consumption = {}
    for meal in meals:
        for ingredient in meal.recipe.ingredients:
            qty = ingredient.quantity * meal.servings / meal.recipe.servings
            key = ingredient.item_id
            
            if key not in consumption:
                consumption[key] = {
                    'item': ingredient.item,
                    'quantity': 0,
                    'unit': ingredient.unit,
                    'total_value': 0
                }
            
            consumption[key]['quantity'] += qty
            if ingredient.item.unit_price:
                consumption[key]['total_value'] += qty * ingredient.item.unit_price
    
    # حساب ملخصات
    total_meals = len(meals)
    total_portions = sum(m.servings for m in meals)
    total_cost = sum(getattr(m.recipe, 'estimated_cost', 0) or 0 for m in meals)
    avg_cost = total_cost / total_meals if total_meals > 0 else 0
    
    # توزيع الوجبات حسب النوع
    meal_breakdown = {
        'breakfast': sum(1 for m in meals if m.meal_type == 'breakfast'),
        'lunch': sum(1 for m in meals if m.meal_type == 'lunch'),
        'dinner': sum(1 for m in meals if m.meal_type == 'dinner'),
        'snack': sum(1 for m in meals if m.meal_type == 'snack')
    }
    
    # أكثر الوصفات استهلاكاً
    recipe_stats = {}
    for meal in meals:
        recipe_id = meal.recipe_id
        if recipe_id not in recipe_stats:
            recipe_stats[recipe_id] = {
                'name': meal.recipe.name,
                'count': 0,
                'total_portions': 0,
                'total_cost': 0
            }
        recipe_stats[recipe_id]['count'] += 1
        recipe_stats[recipe_id]['total_portions'] += meal.servings
        recipe_stats[recipe_id]['total_cost'] += getattr(meal.recipe, 'estimated_cost', 0) or 0
    
    top_recipes = sorted(recipe_stats.values(), key=lambda x: x['count'], reverse=True)[:5]
    
    org_settings = OrganizationSettings.query.first()
    
    return render_template(
        'reports/meal_consumption.html',
        meals=meals,
        consumption=consumption.values(),
        total_meals=total_meals,
        total_portions=total_portions,
        total_cost=total_cost,
        avg_cost=avg_cost,
        meal_breakdown=meal_breakdown,
        top_recipes=top_recipes,
        from_date=from_date,
        to_date=to_date,
        org_settings=org_settings
    )

# ==================== Export Functions ====================

@reports_bp.route('/export/inventory-movement-pdf')
@login_required
def export_inventory_movement_pdf():
    """تصدير تقرير حركة المخزون إلى PDF"""
    from_date = request.args.get('from_date')
    to_date = request.args.get('to_date')
    
    if not from_date or not to_date:
        from_date = (date.today() - timedelta(days=30)).isoformat()
        to_date = date.today().isoformat()
    
    from_date_dt = datetime.strptime(from_date, '%Y-%m-%d')
    to_date_dt = datetime.strptime(to_date, '%Y-%m-%d')
    to_date_dt = to_date_dt.replace(hour=23, minute=59, second=59)
    
    transactions = Transaction.query.filter(
        Transaction.transaction_date.between(from_date_dt, to_date_dt)
    ).order_by(Transaction.transaction_date).all()
    
    org_settings = OrganizationSettings.query.first()
    
    # إنشاء PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=1*cm, leftMargin=1*cm,
                           topMargin=1*cm, bottomMargin=1*cm)
    
    styles = getSampleStyleSheet()
    story = []
    
    # رأس الوثيقة
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=14,
        textColor=colors.HexColor('#1a3a52'),
        spaceAfter=12,
        alignment=1  # Center
    )
    
    if org_settings:
        story.append(Paragraph(org_settings.ministry_name, title_style))
        story.append(Paragraph(org_settings.directorate_name, title_style))
        story.append(Paragraph(org_settings.institution_name, title_style))
        story.append(Spacer(1, 0.5*cm))
    
    story.append(Paragraph(f"تقرير حركة المخزون<br/>من {from_date} إلى {to_date}", title_style))
    story.append(Spacer(1, 0.3*cm))
    
    # الجدول
    data = [['التاريخ', 'المرجع', 'الصنف', 'النوع', 'الكمية', 'السعر', 'القيمة']]
    
    for transaction in transactions:
        data.append([
            transaction.transaction_date.strftime('%Y-%m-%d %H:%M'),
            transaction.reference_number,
            transaction.item.name if transaction.item else '-',
            transaction.transaction_type,
            str(transaction.quantity),
            str(transaction.unit_price or 0),
            str(transaction.total_value or 0)
        ])
    
    table = Table(data, colWidths=[1.5*cm, 2*cm, 3*cm, 1.5*cm, 1.2*cm, 1.2*cm, 1.5*cm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a3a52')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(table)
    doc.build(story)
    
    buffer.seek(0)
    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'inventory_movement_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    )

@reports_bp.route('/export/low-stock-excel')
@login_required
def export_low_stock_excel():
    """تصدير تقرير الأصناف منخفضة المخزون إلى Excel"""
    low_stock_items = Item.query.filter(
        Item.quantity_in_stock <= Item.minimum_quantity,
        Item.is_active == True
    ).all()
    
    org_settings = OrganizationSettings.query.first()
    
    # إنشاء Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'الأصناف منخفضة المخزون'
    
    # رأس الوثيقة
    ws['A1'] = org_settings.ministry_name if org_settings else ''
    ws['A2'] = org_settings.directorate_name if org_settings else ''
    ws['A3'] = org_settings.institution_name if org_settings else ''
    ws['A4'] = f'تقرير الأصناف منخفضة المخزون - {datetime.now().strftime("%Y-%m-%d")}'
    
    # تنسيق الرأس
    for row in range(1, 5):
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'A{row}'].alignment = Alignment(horizontal='right')
    
    # رؤوس الأعمدة
    headers = ['الكود', 'الاسم', 'الفئة', 'الكمية الحالية', 'الحد الأدنى', 'الفرق', 'الوحدة']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col)
        cell.value = header
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='1a3a52', end_color='1a3a52', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    # البيانات
    for row, item in enumerate(low_stock_items, 7):
        ws.cell(row=row, column=1).value = item.code
        ws.cell(row=row, column=2).value = item.name
        ws.cell(row=row, column=3).value = item.category.name if item.category else ''
        ws.cell(row=row, column=4).value = item.quantity_in_stock
        ws.cell(row=row, column=5).value = item.minimum_quantity
        ws.cell(row=row, column=6).value = item.minimum_quantity - item.quantity_in_stock
        ws.cell(row=row, column=7).value = item.unit
    
    # تعديل عرض الأعمدة
    for col in range(1, 8):
        ws.column_dimensions[chr(64 + col)].width = 20
    
    # الحفظ
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return send_file(
        buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'low_stock_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )